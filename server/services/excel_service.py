from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List
from models import Student
import os
from datetime import datetime


class ExcelService:
    """Service for exporting student data to Excel."""
    
    @staticmethod
    def export_students(students: List[Student], output_path: str = None) -> str:
        """
        Export list of students to Excel file.
        
        Args:
            students: List of Student model instances
            output_path: Optional custom output path
            
        Returns:
            Path to the generated Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Define headers
        headers = [
            'ID',
            'Student ID',
            'Full Name',
            'Email',
            'Phone',
            'Department',
            'Program',
            'Year of Study',
            'Document Type',
            'Created At',
            'Updated At'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1, value=student.id)
            ws.cell(row=row_num, column=2, value=student.student_id)
            ws.cell(row=row_num, column=3, value=student.full_name)
            ws.cell(row=row_num, column=4, value=student.email)
            ws.cell(row=row_num, column=5, value=student.phone)
            ws.cell(row=row_num, column=6, value=student.department)
            ws.cell(row=row_num, column=7, value=student.program)
            ws.cell(row=row_num, column=8, value=student.year_of_study)
            ws.cell(row=row_num, column=9, value=student.document_type)
            ws.cell(row=row_num, column=10, value=student.created_at.strftime("%Y-%m-%d %H:%M:%S") if student.created_at else "")
            ws.cell(row=row_num, column=11, value=student.updated_at.strftime("%Y-%m-%d %H:%M:%S") if student.updated_at else "")
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Generate filename if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"students_export_{timestamp}.xlsx"
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
